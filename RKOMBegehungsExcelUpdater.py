import pandas as pd
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog as fd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from tkinterdnd2 import TkinterDnD, DND_FILES
from planio.planio_queries import (
    get_begehungsdaten
)

api_key = "547fa8685b2e0a692cdeb33aeda305edd27b0155"

def main():
    # def handle_boolean(value):
    #     if isinstance(value, bool):  # If it's already a boolean, return it as is
    #         return value
    #     elif isinstance(value, str):  # If it's a string
    #         # Convert to lowercase and check if it's 'true' or 'false'
    #         value = value.strip().lower()  # Strip to remove extra spaces and convert to lowercase
    #         if value == 'wahr':
    #             return True
    #         elif value == 'falsch':
    #             return False
    def select_file():
        output_label.configure(text="",bg="white")
        file_label.configure(bg="white")
        filename.configure(bg="white")
        root.configure(bg="white")
        filename.configure(text=fd.askopenfilename())
    def update_file(filename):
        # Function to copy the formatting from one column to another
        

        if filename.cget("text").endswith(".xlsx"):
            output_label.configure(text="Fehler (Zieldatei muss geschlossen sein)",bg="orange")
            file_label.configure(bg="orange")
            filename.configure(bg="orange")
            root.configure(bg="orange")
            begehungsdaten = get_begehungsdaten(api_key,'131')
            # print("planio-------------------------------------")
            # print(begehungsdaten)
            df = pd.read_excel(filename.cget("text"))
            # print("excel--------------------------------------")
            # print(df)
            # Iterate through each row
            for index, row in df.iterrows():
                # You can access individual values by column name
                if(not pd.isnull(row['Strasse'])):
                    zusatz = ""
                    if(not pd.isnull(row['Hhnrzusatz'])):
                        zusatz = row['Hhnrzusatz']
                    adressen = [(str(row['Gfrgebaeudeid']) + " - " + row['Strasse'] + " " +str(int(row['Hhnr'])) + zusatz + ", " + str(int(row['Plz'])) + " " + row['Ort']),(str(row['Gfrgebaeudeid']) + " - " + row['Strasse'] + " " +str(int(row['Hhnr'])) + zusatz + " " + str(int(row['Plz'])) + " " + row['Ort']),(str(row['Gfrgebaeudeid']) + " - " + row['Strasse'] + " " +str(int(row['Hhnr'])) + " " + zusatz + ", " + str(int(row['Plz'])) + " " + row['Ort']),(str(row['Gfrgebaeudeid']) + " - " + row['Strasse'] + " " +str(int(row['Hhnr'])) + " " + zusatz + " " + str(int(row['Plz'])) + " " + row['Ort'])]
                    for adresse in adressen:
                        if not begehungsdaten[begehungsdaten["address"] == adresse].head(1).empty:
                            planio_row = begehungsdaten[begehungsdaten["address"] == adresse].head(1)
                    
                    if not planio_row.empty:
                        df.loc[index, "Status"] = planio_row["status"].iloc[0]
                        if not pd.isnull(planio_row["protokoll"].iloc[0]):
                            df.loc[index, "Potokoll versandt"] = planio_row["protokoll"].iloc[0]
                        df.loc[index, "Abgelegt & Übergeben"] = planio_row["closed_on"].iloc[0]
                        df.loc[index, "Sachstand"] = planio_row["sachstand"].iloc[0]
                        if not pd.isnull(row["Erschließung-Bemerkung"]) and not pd.isnull(planio_row["bemerkung"].iloc[0]):
                            patterns = [r"\nOrtstermin: \d{4}-\d{2}-\d{2}",r"\n\d{1}. Kontaktversuch: \d{4}-\d{2}-\d{2}",r"Ortstermin: \d{4}-\d{2}-\d{2}",r"\d{1}. Kontaktversuch: \d{4}-\d{2}-\d{2}"]
                            # Loop through the patterns
                            for pattern in patterns:
                                # Use re.sub() to replace matched patterns with an empty string
                                row["Erschließung-Bemerkung"] = re.sub(pattern, "", row["Erschließung-Bemerkung"])

                            # Clean up leading/trailing spaces after replacement
                            row["Erschließung-Bemerkung"] = row["Erschließung-Bemerkung"].strip()
                            print(row["Erschließung-Bemerkung"])
                            df.loc[index, "Erschließung-Bemerkung"] = row["Erschließung-Bemerkung"] + "\n" + planio_row["bemerkung"].iloc[0]
                        elif not pd.isnull(planio_row["bemerkung"].iloc[0]):
                            df.loc[index, "Erschließung-Bemerkung"] = "\n" + planio_row["bemerkung"].iloc[0]
                        else:
                            df.loc[index, "Erschließung-Bemerkung"] = row["Erschließung-Bemerkung"]
                        # if handle_boolean(row["Nutzungsvereinbarung"]) and planio_row["status"].iloc[0] == "Wartend":
                        #     print("change" + str(planio_row["issue_id"].iloc[0]))
                            


            # Load the existing workbook
            book = load_workbook(filename.cget("text"))
            sheet = book.worksheets[0]
            def copy_column_formatting(sheet, source_col, target_col, start_row=1, end_row=sheet.max_row):
                for row in range(start_row, end_row + 1):
                    # Copy the cell formatting from the source column to the target column
                    source_cell = sheet.cell(row=row, column=source_col)
                    target_cell = sheet.cell(row=row, column=target_col)

                    # Copy font
                    target_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size, 
                                bold=source_cell.font.bold, italic=source_cell.font.italic, 
                                color=source_cell.font.color, underline=source_cell.font.underline)
                    # Copy fill (background color)
                    if source_cell.fill.start_color:
                        target_cell.fill = PatternFill(start_color=source_cell.fill.start_color, 
                                                    end_color=source_cell.fill.end_color, 
                                                    fill_type=source_cell.fill.fill_type)
                    else:
                        target_cell.fill = PatternFill(fill_type='none')  # If no fill, apply no fill
                    # Copy alignment
                    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                           vertical=source_cell.alignment.vertical,
                                           text_rotation=source_cell.alignment.text_rotation,
                                           wrap_text=source_cell.alignment.wrap_text,
                                           shrink_to_fit=source_cell.alignment.shrink_to_fit)
                    # Copy borders
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom
                    )

            # Function to adjust row height based on wrapped text
            def adjust_row_height(sheet, row_num, min_height=25):
                max_length = 0
                for col_num in range(1, 18):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Estimate row height based on max length of the content
                # You can adjust this heuristic value as necessary (e.g., the multiplier `1.2` may need tweaking).
                estimated_height = max(min_height, int(max_length / 2) * 1.2)  # Adjust multiplier as needed
                sheet.row_dimensions[row_num].height = estimated_height

            # Function to adjust column widths to the longest string, excluding the header
            def adjust_columns_to_longest_line(sheet):
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)

                    # Skip the first row (header)
                    for cell in col[1:]:  # Start from the second row (index 1)
                        if cell.value:
                            # Consider line breaks, so split the text by line breaks and calculate the longest line
                            lines = str(cell.value).split('\n')
                            max_length = max(max_length, max(len(line) for line in lines))

                    # Adjust the width by adding padding (you can adjust this number as needed)
                    adjusted_width = max_length + 2  # Add some padding
                    sheet.column_dimensions[column].width = adjusted_width

            # Specify the column range you want to copy formatting from (for example, columns 1 to 4)
            start_col = 1  # Column A
            end_col = 26    # Column D (you can adjust this)
            new_start_col = 2 # Start from column E to extend formatting

            for col in range(start_col, end_col + 1):
                copy_column_formatting(sheet, start_col, col + (new_start_col - start_col))
            # Write the column names (headers) to the first row
            for col_num, column_name in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col_num, value=column_name)

            # Write the data to the sheet, starting from row 2
            for row_num, row in enumerate(df.values, start=2):
                for col_num, value in enumerate(row, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)
                adjust_row_height(sheet,row_num)
            adjust_columns_to_longest_line(sheet)
            # Set column widths
            sheet.column_dimensions['I'].width = 20
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 20
            sheet.column_dimensions['O'].width = 25
            # Specify the column you want to modify (e.g., column 'A')
            column_letter = 'S'

            # Loop through all rows in the specified column and set text alignment to left
            for row in sheet.iter_rows(min_col=sheet[column_letter][0].column, max_col=sheet[column_letter][0].column, min_row=1, max_row=sheet.max_row):
                for cell in row:
                    # Set the cell's horizontal alignment to left
                    cell.alignment = Alignment(horizontal='left')

            # Save the modified workbook
            book.save(filename.cget("text"))

            output_label.configure(text="Fertig",bg="lightgreen")
            file_label.configure(bg="lightgreen")
            filename.configure(text="",bg="lightgreen")
            root.configure(bg="lightgreen")
        else:
            output_label.configure(text="Datei nicht kompatibel",bg="yellow")
            file_label.configure(bg="yellow")
            filename.configure(bg="yellow")
            root.configure(bg="yellow")

    def on_resize(event):
        file_label.configure(wraplength=root.winfo_width()-50)
        output_label.configure(wraplength=root.winfo_width()-50)
        filename.configure(wraplength=root.winfo_width()-50)
    def on_drop(event):
        output_label.configure(text="",bg="white")
        file_label.configure(bg="white")
        filename.configure(bg="white")
        root.configure(bg="white")
        # Get the file path from the event and display it
        file_path = event.data
        file_name = file_path.split("{")[1] 
        file_name = file_name.split("}")[0]  
        filename.config(text=file_name)
    # Create the main window
    root = TkinterDnD.Tk()
    window_width = 400
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int((screen_height - window_height) / 2)
    position_left = int((screen_width - window_width) / 2)
    root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")
    root.configure(bg="white")
    root.title("RKOM Begehungs Excel Updater")
    # Create a label widget
    root.bind("<Configure>",on_resize)
    file_label = tk.Label(root, text='Gewählte Datei:',bg="white",wraplength = window_width-50)
    file_label.pack()
    filename = tk.Label(root,text="Keine",bg="white",wraplength = window_width-50) #Keine
    filename.pack()
    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', on_drop)
    file_button = ttk.Button(root,text="Datei auswählen",command = lambda: select_file())
    file_button.pack()
    update_button = ttk.Button(root,text="Updaten",command = lambda: update_file(filename))
    update_button.pack()
    
    output_label = tk.Label(root, text='',bg="white",wraplength = window_width-50)
    output_label.pack()
    # Start the GUI event loop
    root.mainloop()

main()